import PySimpleGUI as sg
import pandas as pd
import numpy as np
import time
import threading
import subprocess
import gspread
from google.oauth2 import service_account
from openpyxl import load_workbook
from concurrent.futures import ThreadPoolExecutor
from pathlib import Path

#Uploading Code to GitHub Test

scopes = ["https://www.googleapis.com/auth/spreadsheets",
          "https://www.googleapis.com/auth/drive"]
json_file = "credentials.json"

sg.theme_add_new(
    'NewTheme38029', 
    {
        'BACKGROUND': '#200f21', 
        'TEXT': '#f638dc', 
        'INPUT': '#5a3d5c', 
        'TEXT_INPUT': '#FFFFFF', 
        'SCROLL': '#5a3d5c', 
        'BUTTON': ('#FFFFFF', '#382039'), 
        'PROGRESS': ('#000000', '#000000'), 
        'BORDER': 0, 
        'SLIDER_DEPTH': 0, 
        'PROGRESS_DEPTH': 0, 
        'COLOR_LIST': ['#200f21', '#382039', '#5a3d5c', '#f638dc'], 
    }
)
sg.theme('NewTheme38029')
ttk_style = 'vista'

month_names = {
    1: "Janeiro",
    2: "Fevereiro",
    3: "Março",
    4: "Abril",
    5: "Maio",
    6: "Junho",
    7: "Julho",
    8: "Agosto",
    9: "Setembro",
    10: "Outubro",
    11: "Novembro",
    12: "Dezembro"
}

def get_month_name(month_number):
    if month_number in month_names:
        return month_names[month_number]
    else:
        return "Invalid month number"  # Handle invalid inputs
    
def create_file_path(folder_path, month_name, result_name):
    return Path(folder_path) / f"{month_name}_{result_name}.xlsx"

def write_to_excel(df, file_old, file_path, append_to_file):
    """
    Write or append a DataFrame to an Excel file.

    Parameters:
    df : DataFrame
        DataFrame to write to the Excel file.
    file_path : str or Path
        Path to the Excel file.
    append_to_file : bool
        True to append to the file, False to overwrite.
    """
    # Check if appending to file
    if append_to_file:
        try:
            # Load the existing workbook
            book = load_workbook(file_old)
            writer = pd.ExcelWriter(file_old, engine='openpyxl')
            writer.book = book
            writer.sheets = {ws.title: ws for ws in book.worksheets}
            
            # Get the last row in the existing sheet
            startrow = book.active.max_row

            # Write the new data after the last row
            df.to_excel(writer, index=False, header=False, startrow=startrow)
            writer.save()
            writer.close()
        except FileNotFoundError:
            # If file does not exist, write a new file
            df.to_excel(file_path, index=False)
    else:
        # Write a new file
        df.to_excel(file_path, index=False)

def open_file(file_path):
    try:
        subprocess.Popen(["start", file_path], shell=True)  # On Windows
        return f"Opening {file_path} in the default application... Success!"
    except FileNotFoundError:
        return f"Error: File '{file_path}' not found."
    except Exception as e:
        return f"An error occurred: {e}"

def login():
    credentials = service_account.Credentials.from_service_account_file(json_file)
    scoped_credentials = credentials.with_scopes(scopes)
    gc = gspread.authorize(scoped_credentials)
    return gc

# Function that performs the main code execution and updates status
def execute_code(file_paths, window, append_to_file, gc):
    # Get the elements of the status window
    status_elem = window['_STATUS_']

    time.sleep(2)
    
    status_elem.print('bip bip, carregando seus arquivos :)')
    # Function to load an Excel file with specific parameters
    def load_excel(file_param):
        file_path = file_param['file_path']
        kwargs = file_param.get('kwargs', {})  # Extracting optional loading parameters
        
        return pd.read_excel(file_path, **kwargs)

    # File paths and respective loading parameters for your Excel files
    file_params = [
        {
            'file_path': f'{file_paths[0]}',
            'kwargs': {'skiprows': 6}
        },
        {
            'file_path': f'{file_paths[1]}'
        }
    ]

    # Initialize a ThreadPoolExecutor
    with ThreadPoolExecutor() as executor:
        # Load Excel files concurrently using threads
        results = list(executor.map(load_excel, file_params))

    # Results will contain loaded DataFrames for each Excel file
    # You can access them from the results list
    df, Cdf = results  # Assign loaded DataFrames to variables

    df = df.iloc[:-1]

    gc = login()
    planilha = gc.open("BASQUETE")
    aba = planilha.worksheet("2024")
    dados = aba.get_all_records()
    Bdf = pd.DataFrame(dados)

    print(Bdf)
    time.sleep(2)
    
    status_elem.print('bip bop bop. Realizando muitos, MUITOS, cálculos')

    time.sleep(3)

    # Strip whitespace while preserving NaN values
    df['Emissora TV'] = df['Emissora TV'].str.strip()
    df['Anunciante'] = df['Anunciante'].str.strip()
    df['Marca'] = df['Marca'].str.strip()
    df['Agência'] = df['Agência'].str.strip()
    df['Marca'] = df['Marca'].str.strip()
    Bdf['Cliente'] = Bdf['Cliente'].str.strip()

    #correcting emissora that came with wrong data
    mask = df['Emissora TV'] == 'RECORD'
    df.loc[mask, 'Emissora TV'] = 'RECORD TV'

    time.sleep(3)

    df['INV(000)'] = df['INV(000)'].astype(np.float32)
    df['Inserção'] = df['Inserção'].astype(np.int16)
    df['Ano-Mês'] = df['Ano-Mês'].astype(np.int32)
    df['Praça'] = df['Praça'].astype('category')
    df['Emissora TV'] = df['Emissora TV'].astype('category')
    df['Categoria'] = df['Categoria'].astype('category')
    df['Tipo Veiculação'] = df['Tipo Veiculação'].astype('category')
    Bdf['Mês'] = pd.to_datetime(Bdf['Mês'])

    print(df.dtypes)

    time.sleep(3)

    conditions = (
        df['Anunciante'].str.contains(r'\bGLOBO\b', case=False, regex=True) |
        df['Anunciante'].str.contains('BANDEIRANTES') |
        df['Anunciante'].str.contains('RECORD') |
        df['Anunciante'].str.contains('GAZETA') |
        df['Anunciante'].str.contains('CNT') |
        df['Anunciante'].str.contains('JOVEM PAN') |
        df['Anunciante'].str.contains('RADIO') |
        df['Anunciante'].str.contains('MASSA') |
        df['Anunciante'].str.contains('TELEVISAO') |
        df['Anunciante'].str.contains('TV') |
        df['Anunciante'].str.contains('SBT') |
        df['Anunciante'].str.contains('RIC') |
        (df['Marca'].str.contains('CARTOLA', na=False)) |
        df['Anunciante'].str.contains('{DESCONHECIDO}') |
        df['Anunciante'].str.contains('TSE') |
        df['Categoria'].str.contains('CAMPANHAS BENEFICIENTES SOCIAIS') |
        df['Categoria'].str.contains('CAMPANHAS PARTIDARIAS') |
        (df['Marca'].str.contains('BEECON')) |
        (df['Marca'].str.contains('TOPVIEW'))
    )

    df.drop(df[conditions].index, inplace=True)

    #function to extract a value from the row (finding the city name from the prefecture name)
    def extract_value(row):
        start = 'PREF MUN '
        end = ' (GMP)'
        value = row['Anunciante']
        if start in value and end in value:
            extracted_value = value.split(start)[1].split(end)[0]
            return extracted_value
        else:
            return row['Cidade Autorização']

    #apply the city name from the correct prefecture to the city column
    df['Cidade Autorização'] = df.apply(extract_value, axis=1)

    #rename the column from the gov est pr to match it's location
    df['UF Autorização'] = np.where(df['Anunciante'].str.contains('GOV EST PR (GEP)', 
                                                                  regex=False), 'PARANA', df['UF Autorização'])
    df['Cidade Autorização'] = np.where(df['Anunciante'].str.contains('GOV EST PR (GEP)', 
                                                                      regex=False), 'CURITIBA', df['Cidade Autorização'])

    print(df)

    #correcting some clients that came with wrong data
    mask = df['Marca'] == 'MUFFATAO'
    df.loc[mask, 'Anunciante'] = 'PEDRO MUFFATO & CIA LTDA'

    clients_maringa = (
        (df['Marca'] == 'CANCAO') |
        df['Marca'].str.contains('AMIGAO', na=False) |
        df['Marca'].str.contains('ZAELI', na=False) |
        ((df['Marca'].str.contains('IMPACTO', na=False)) & (df['Praça'].str.contains('MARINGA', na=False))) |
        ((df['Marca'].str.contains('SOLUCIONADOR', na=False)) & (df['Praça'].str.contains('MARINGA', na=False))) |
        df['Marca'].str.contains('COAMO', na=False) |
        df['Marca'].str.contains('UNICESUMAR', na=False) |
        df['Marca'].str.contains('ORAL SIN', na=False)
    )
    clients_curitiba = (
        df['Marca'].str.contains('PEDROSO', na=False) |
        (df['Anunciante'] == 'ALTHAIA') |
        (df['Marca'].str.contains('ORAL UNIC', na=False)) & (df['Praça'].str.contains('CURITIBA', na=False)) |
        (df['Anunciante'] == 'LIGGA TELECOM') |
        df['Marca'].str.contains('CARRERA CARNEIRO', na=False) |
        (df['Marca'].str.contains('ORAL SIN', na=False)) & (df['Praça'].str.contains('CURITIBA', na=False))
    )
    clients_cascavel = (
        (df['Marca'].str.contains('IMPACTO PRIME', na=False)) & (df['Praça'].str.contains('CASCAVEL', na=False)) |    
        (df['Marca'].str.contains('SUPERGASBRAS ENERGIA', na=False)) & (df['Praça'].str.contains('CASCAVEL', na=False)) |
        df['Marca'].str.contains('ARMAZEM DA MARIA', na=False) |
        (df['Marca'].str.contains('BLUEFIT', na=False)) & (df['Praça'].str.contains('CASCAVEL', na=False)) |
        df['Anunciante'].str.contains('CBS EMPREENDIMENTOS IMOBILIARIOS', na=False) |
        (df['Marca'].str.contains('UMUPREV', na=False)) & (df['Praça'].str.contains('CASCAVEL', na=False)) |
        (df['Marca'].str.contains('ORAL UNIC', na=False)) & (df['Praça'].str.contains('CASCAVEL', na=False)) |
        df['Marca'].str.contains('SHOPPING CHINA', na=False) |
        df['Marca'].str.contains('ODONTO SAN', na=False) |
        (df['Anunciante'].str.contains('CIA BEAL DE ALIMENTOS', na=False)) & (df['Praça'].str.contains('CASCAVEL', na=False)) |
        df['Anunciante'].str.contains('IND E COM DE LATICINIOS PEREIRA', na=False) |
        (df['Marca'] == 'MUFFATAO') |
        (df['Marca'] == 'ITAIPU BINACIONAL') |
        (df['Marca'] == 'FOZ TINTAS') |
        (df['Marca'] == 'UNIPRIME')
    )
    clients_londrina = (
        df['Anunciante'].str.contains('SUPER MUFFATO', na=False) |
        df['Marca'].str.contains('SOLUCAO', na=False)
    )

    # Rename the city column that commonly mismatches its real location
    df.loc[clients_maringa, ['Cidade Autorização', 'UF Autorização']] = ['MARINGA', 'PARANA']
    df.loc[clients_curitiba, ['Cidade Autorização', 'UF Autorização']] = ['CURITIBA', 'PARANA']
    df.loc[clients_londrina, ['Cidade Autorização', 'UF Autorização']] = ['LONDRINA', 'PARANA']
    df.loc[clients_cascavel, ['Cidade Autorização', 'UF Autorização']] = ['CASCAVEL', 'PARANA']

    mask = df['Marca'] == 'MAX' 
    df.loc[mask, 'Agência'] = 'NOBRE PROPAGANDA'

    # # New columns and filling those columns

    #creates a few columns
    df['Vl Tab (000)'] = df['INV(000)'] * 1000
    df['Desconto'] = 0
    df['Valor Líquido Projetado'] = 0
    df['Cobertura'] = None
    df['Região'] = None
    df['Mercado'] = None
    df['Data'] = pd.to_datetime(df['Ano-Mês'], format='%Y%m')

    df.reset_index(inplace=True, drop=True)

    status_elem.print('ensinando marmotas a dançarem macarena')

    #function to determine coverage:
    def determine_coverage(row, Cdf):
        cidade_autorizacao = row['Cidade Autorização'].upper()
        
        if 'COBERTURA' in Cdf.columns and pd.notna(Cdf.loc[Cdf['Municipio'] == cidade_autorizacao, 'COBERTURA']).any():
            return Cdf.loc[Cdf['Municipio'] == cidade_autorizacao, 'COBERTURA'].values[0]
        else:
            return 'IMPORT'
        
    #function to determine our coverage region:
    cidade_region_map = dict(zip(Cdf['Municipio'], Cdf['Região']))

    #function to determine our market:
    def set_market(row):
        UF = row['UF Autorização']
        anunciante = row['Anunciante'].upper()
        cidade_autorizacao = row['Cidade Autorização'].upper()

        if 'PREF' in anunciante and (cidade_autorizacao in ['CURITIBA', 'MARINGA', 'CASCAVEL', 
                                                            'TOLEDO', 'FOZ DO IGUACU', 'LONDRINA']):
            return 'PREF SEDE'
        elif 'PREF' in anunciante:
            return 'PREF'
        elif 'GOV' in anunciante and 'FEDERAL' not in anunciante:
            return 'GOVERNO'
        elif '(GEP)' in anunciante and ('ASSEMBLEIA' not in anunciante):
            return 'GOVERNO'
        elif 'ASSEMBLEIA' in anunciante:
            return 'ASSEMBLEIA'
        
        return 'LOCAL' if UF == 'PARANA' else 'IMPORT'

    #fill some of our Columns:
    df['Cobertura'] = df.apply(determine_coverage, args=(Cdf,), axis=1)
    df['Região'] = df['Cidade Autorização'].map(cidade_region_map).fillna('IMPORT')
    df['Mercado'] = df.apply(set_market, axis=1)

    status_elem.print('marmotas dançam macarena também')

    df['Mercado'] = df['Mercado'].astype('category')
    df['Cobertura'] = df['Cobertura'].astype('category')
    df['Região'] = df['Região'].astype('category')

    print(df.dtypes)
    print(df)

    # #### Setting our discounts

    def discount_giver(row):
        emissora = row['Emissora TV']
        coverage = row['Cobertura']
        praca = row['Praça']
        anunciante = row['Anunciante']
        
        #SBT discounts
        if 'SBT' in emissora and 'MERCHANDISING' not in praca:
            return 0.94
        elif 'SBT' in emissora and 'MERCHANDISING' in praca:
            return 0.93
        #BAND discounts
        elif 'BANDEIRANTES' in emissora and 'MERCHANDISING' not in praca and (praca in ['LONDRINA', 'MARINGA', 'CURITIBA']):
            return 0.95
        elif 'BANDEIRANTES' in emissora and 'MERCHANDISING' in praca and (praca in ['LONDRINA', 'MARINGA', 'CURITIBA']):
            return 0.95
        elif 'BANDEIRANTES' in emissora and 'MERCHANDISING' not in praca and 'CASCAVEL' in praca:
            return 0.95
        elif 'BANDEIRANTES' in emissora and 'MERCHANDISING' in praca and 'CASCAVEL' in praca:
            return 0.95
        #CNT discounts
        elif 'CNT' in emissora and 'CURITIBA' in praca:
            return 0.9
        #GLOBO discounts
        elif 'GLOBO' in emissora and 'MERCHANDISING' not in praca and (praca in ['MARINGA', 'LONDRINA', 'FOZ DO IGUACU']):
            return 0.3
        elif 'GLOBO' in emissora and 'MERCHANDISING' in praca and (praca in ['MARINGA', 'LONDRINA', 'FOZ DO IGUACU']):
            return 0.29
        elif 'GLOBO' in emissora and 'MERCHANDISING' not in praca and (praca in ['PARANAVAI', 'PONTA GROSSA', 'GUARAPUAVA']):
            return 0.4
        elif 'GLOBO' in emissora and 'MERCHANDISING' not in praca and (praca in ['CURITIBA', 'CASCAVEL']):
            return 0.3
        elif 'GLOBO' in emissora and 'MERCHANDISING' in praca and (praca in ['CURITIBA', 'CASCAVEL']):
            return 0.24
        
        #Import discounts
        elif 'BANDEIRANTES' in emissora and 'IMPORT' in coverage:
            return 0.95
        elif 'SBT' in emissora and 'IMPORT' in coverage:
            return 0.9
        elif 'GLOBO' in emissora and 'IMPORT' in coverage:
            return 0.1
        elif 'CNT' in emissora and 'IMPORT' in coverage:
            return 0.96
        elif 'RECORD' in emissora and 'IMPORT' in coverage:
            return 0.82
        
        #Specific Discounts
        #SBT discounts:
        elif 'SBT' in emissora and 'CONDOR' in anunciante and 'MERCHANDISING' not in praca:
            return 0.89
        elif 'SBT' in emissora and 'CONDOR' in anunciante and 'MERCHANDISING' in praca:
            return 0.88
        elif 'SBT' in emissora and 'SUPER MUFFATO' in anunciante and 'MERCHANDISING' not in praca:
            return 0.89
        elif 'SBT' in emissora and 'SUPER MUFFATO' in anunciante and 'MERCHANDISING' in praca:
            return 0.88
        elif 'SBT' in emissora and 'ALIANCA' in anunciante and 'MERCHANDISING' not in praca:
            return 0.9950
        elif 'SBT' in emissora and 'ALIANCA' in anunciante and 'MERCHANDISING' in praca:
            return 0.93
        elif 'SBT' in emissora and 'RIO VERDE' in anunciante:
            return 0.98
        elif 'SBT' in emissora and 'ODONTO EXCELLENCE' in anunciante:
            return 0.96
        #BAND discounts
        elif 'BANDEIRANTES' in emissora and 'MALUCELLLI' in anunciante:
            return 0.98
        elif 'BANDEIRANTES' in emissora and 'PONTO DE VISAO' in anunciante:
            return 0.9850
        elif 'BANDEIRANTES' in emissora and 'O SOLUCIONADOR' in anunciante:
            return 0.98
        elif 'BANDEIRANTES' in emissora and 'SUPER MUFFATO' in anunciante:
            return 0.97
        #GLOBO discounts
        elif 'GLOBO' in emissora and 'COORITIBA FOOT BALL CLUB' in anunciante:
            return 0.5
        elif 'GLOBO' in emissora and 'PONTO DE VISÃO' in anunciante:
            return 0.7
        elif 'GLOBO' in emissora and 'KURTEN' in anunciante:
            return 0.55
        elif 'GLOBO' in emissora and 'JOCKEY PLAZA SHOP' in anunciante:
            return 0.5
        #GOV discounts
        elif 'CNT' not in emissora and 'RECORD' not in emissora and 'GOV' in anunciante:
            return 0.13
        elif 'CNT' in emissora and 'GOV' in anunciante:
            return 0.50
        #ASSEM discounts
        elif 'CNT' not in emissora and 'RECORD' not in emissora and 'ASSEMBLEIA' in anunciante:
            return 0.13
        elif 'CNT' in emissora and 'ASSEMBLEIA' in anunciante:
            return 0.50
        #PREF discounts
        elif 'GLOBO' in emissora and 'PREF MUN CURITIBA (GMP)' in anunciante:
            return 0.15
        elif 'BANDEIRANTES' in emissora and 'PREF MUN CURITIBA (GMP)' in anunciante:
            return 0.20
        elif 'SBT' in emissora and 'PREF MUN CURITIBA (GMP)' in anunciante:
            return 0.20
        elif 'CNT' in emissora and 'PREF MUN CURITIBA (GMP)' in anunciante:
            return 0.55
        else:
            return 0

    df['Desconto'] = df.apply(discount_giver, axis=1)

    # Apply some discounts
    mask1 = (df['Desconto'] != 0) & (df['Agência'] == '{DIRETO}')  # Check for direct agency
    mask2 = (df['Desconto'] != 0) & (df['Agência'] != '{DIRETO}')  # Check for non-direct agency

    df['Valor Líquido Projetado'] = 0  # Initialize the column with zeros

    df.loc[mask1, 'Valor Líquido Projetado'] = df['Vl Tab (000)'] * (1 - df['Desconto'])
    df.loc[mask2, 'Valor Líquido Projetado'] = df['Vl Tab (000)'] * (1 - df['Desconto']) * (1 - 0.2)

    #Filters the Bdf to better manage memory:
    Bdf['Mês'] = pd.to_datetime(Bdf['Mês'], format='%m-%Y')
    condition1 = Bdf['Mês'].dt.month == int(file_paths[2])
    condition2 = Bdf['Exibição'].isin(['CURITIBA', 'LONDRINA', 'TOLEDO', 'MARINGÁ'])
    condition3 = Bdf['Contato'] != 'PERMUTA'

    Bdf = Bdf[condition1 & condition2 & condition3]

    print(Bdf)

    #Updates values from Governo and Assembleia contacts
    def update_valor_liquido_based_on_contato(row, Bdf, df):
        # Define a mapping for conditions based on praca, exibicao, and contato in Bdf
        conditions = {
            ('CASCAVEL', 'TOLEDO', 'GOVERNO'): 'GOV',
            ('CASCAVEL', 'TOLEDO', 'ASSEMBLEIA'): 'ASSEMBLEIA',
            ('CURITIBA', 'CURITIBA', 'GOVERNO') : 'GOV',
            ('CURITIBA', 'CURITIBA', 'ASSEMBLEIA') : 'ASSEMBLEIA',
            ('LONDRINA', 'LONDRINA', 'GOVERNO') : 'GOV',
            ('LONDRINA', 'LONDRINA', 'ASSEMBLEIA') : 'ASSEMBLEIA',
            ('MARINGA', 'MARINGÁ', 'GOVERNO') : 'GOV',
            ('MARINGA', 'MARINGÁ', 'ASSEMBLEIA') : 'ASSEMBLEIA'
        }
        
        anunciante = row['Anunciante'].upper()
        praca = row['Praça'].upper()
        emissora = row['Emissora TV'].upper()
        month = row['Data'].month

        # Iterate through conditions to find a match
        for (praca_cond, exibicao_cond, contato_cond), anunciante_cond in conditions.items():
            if anunciante_cond in anunciante and praca_cond in praca and 'RECORD' in emissora:
                filtered_rows = Bdf[(Bdf['Contato'] == contato_cond) & (Bdf['Exibição'].str.contains(exibicao_cond))]

                sum_value = filtered_rows['Valor Líquido'].sum()

                # Debug print statements
                print(f"Row Index: {row.name}")
                print(f"Filtered Rows:\n{filtered_rows}")

                # Create a mask for the filtered rows
                mask = (df['Anunciante'].str.upper().str.contains(anunciante_cond)) & \
                    (df['Praça'].str.upper().str.contains(praca_cond)) & \
                    (df['Emissora TV'].str.upper().str.contains('RECORD'))

                # Count the number of rows that meet the filter criteria in your main DataFrame (df)
                num_filtered_rows_in_df = len(df[mask])

                if num_filtered_rows_in_df > 0:
                    # Distribute the sum_value equally among the filtered rows in your main DataFrame
                    value_per_row = sum_value / num_filtered_rows_in_df
                    
                    # Update the 'Valor Líquido Projetado' column for the filtered rows
                    return value_per_row

        return row['Valor Líquido Projetado']

    # Apply the function to each row
    df['Valor Líquido Projetado'] = df.apply(lambda row: update_valor_liquido_based_on_contato(row, Bdf, df), axis=1)

    status_elem.print('blip blip. Extraindo valores de CURITIBA e colocando no seu relatório')

    #This is a Dictionary that manually corrects mismatched values based on our findings in the Bdf:
    replace_dict = {
        'TERRITORIO DA AGUIA COMERCIO DE CALCADOS - EIRELI - EPP': 'SERALLE COMERCIO DE CALCADOS LTDA',
        'CASTARDO COMERCIAL DE CALÇADOS LTDA': 'SERALLE COMERCIO DE CALCADOS LTDA',
        'O SOLUCIONADOR LONDRINA ASSESSORIA LTDA' : 'O SOLUCIONADOR ASSESSORIA LTDA',
        'O SOLU MARINGA ASSESSRIA FINANCEIRO LTDA' : 'O SOLUCIONADOR ASSESSORIA LTDA',
        'O SOLUCIONADOR CURITIBA ASSESSORIA LTDA' : 'O SOLUCIONADOR ASSESSORIA LTDA',
        'IRMAOS MUFFATO S.A' : 'IRMAOS MUFFATO CIA LTDA',
        'LOVAT VEICULOS S/A' : 'LOVAT VEÍCULOS LTDA',
        'ITALO SUPERMERCADO LTDA' : 'ITALO SUPERMERCADOS LTDA',
        'COOP DE C E I DE L AD V DA REG DAS CAT DO IG E VALE DO PAR' : 'C.VALE COOPERATIVA AGROINDUSTRIAL',
        'COOPERATIVA AGROP. MOURAOENSE LTDA.' : 'COAMO COOPERATIVA AGROINDUSTRIAL',
        'ASSESSORIA EXTRAJUDICIAL SOLUCAO FINANCEIRA EIRELI' : 'SOLUCAO FINANCEIRA - SERVICOS DE RECUPERACAO DE CREDITO EIRELI',
        'EAA - TOLEDO COMERCIO DE PECAS E PNEUS LTDA' : 'J A L IMPACTO MARINGA COMERCIO DE PECAS E PNEUS EIRELI',
        'ODONTOTOP FRANCISCO BELTRAO LTDA' : 'ODONTOTOP TOLEDO LTDA',
        'ODONTOTOP MARECHAL CANDIDO RONDON LTDA' : 'ODONTOTOP TOLEDO LTDA',
        'ODONTOTOP REALEZA LTDA' : 'ODONTOTOP TOLEDO LTDA',
        'ODONTOTOP CASCAVEL LTDA' : 'ODONTOTOP TOLEDO LTDA',
        'UMUPREV PLANO ASSISTENCIA FAMILIAR LTDA' : 'UMUPREV PLANO DE ASSISTENCIA FAMILIAR LTDA',
        'ORAL UNIC ODONTOLOGIA ARAUCARIA LTDA' : 'ORAL UNIC ODONTOLOGIA TOLEDO LTDA',
        'ORAL UNIC ODONTOLOGIA CAMPO LARGO LTDA' : 'ORAL UNIC ODONTOLOGIA TOLEDO LTDA',
        'ORAL UNIC ODONTOLOGIA ALMIRANTE TAMANDARE LTDA' : 'ORAL UNIC ODONTOLOGIA TOLEDO LTDA',
        'ORAL UNIC ODONTOLOGIA PINHAIS LTDA' : 'ORAL UNIC ODONTOLOGIA TOLEDO LTDA',
        'ORAL UNIC ODONTOLOGIA FAZENDA RIO GRANDE LTDA' : 'ORAL UNIC ODONTOLOGIA TOLEDO LTDA',
        'ORAL UNIC ODONTOLOGIA PINHAIS LTDA' : 'ORAL UNIC ODONTOLOGIA TOLEDO LTDA',
        'ORAL UNIC ODONTOLOGIA COLOMBO LTDA' : 'ORAL UNIC ODONTOLOGIA TOLEDO LTDA',
        'GCA DISTRIBUIDORA COMERCIAL DE ALIMENTOS LTDA' : 'GCA - DISTRIBUIDORA COMERCIAL DE ALIMENTOS LTDA',
    }

    # Use the replace method to replace values in the 'Cliente' column
    Bdf['Cliente'].replace(replace_dict, inplace=True)


    #now we know that some or most of our clients might have different names than what it's in our df 
    client_mapping = {
        'SERALLE CALCADOS' : 'SERALLE COMERCIO DE CALCADOS LTDA',
        'OTICA COMERCIAL' : 'ESCORPIAO JOIAS E RELOGIOS LTDA',
        'CASTRO E ROSA' : 'CASTRO E ROSA LTDA',
        'LOJAS MM MERCADOMOVEIS' : 'MERCADOMOVEIS LTDA',
        'ASSOC DOS LOJISTAS DO PALLADIUM SHOPPING CENTER' : 'ASSOCIAÇÃO DOS LOJISTAS DO PALLADIUM SHOPPING CENTER',
        'ASSOC PARA O FUNDO DE PROMOCAO DO VENTURA SHOPPING' : 'ASSOCIACAO PARA O FUNDO DE PROMOÇÃO DE VENTURA SHOPPING DE DESCONTOS',
        'BARIGUI VEICULOS' : 'AUTOBARIGUI COMERCIO DE VEICULOS LTDA',
        'TOKO FRIOS ALIMENTOS' : 'TOKO FRIOS ALIMENTOS LTDA.ME',
        'DAJU' : 'DAJU COMERCIO DE TECIDOS LTDA',
        'CONDOR SUPER CENTER' : 'CONDOR SUPER CENTER LTDA',
        'GIGANTE ATACADISTA' : 'GARANTE DISTRIB E IMPORT DE PROD ALIMENTICIOS LTDA',
        'MOVEIS CAMPO LARGO IND E COM' : 'MOVEIS CAMPO LARGO INDUSTRIA E COMERCIO  LTDA',
        'UNITECH IND DE MAQUINAS' : 'UNITECH INDUSTRIA DE MAQUINAS LTDA',
        'O SOLUCIONADOR ASSESSORIA' : 'O SOLUCIONADOR ASSESSORIA LTDA',
        'SUPER MUFFATO' : 'IRMAOS MUFFATO CIA LTDA',
        'FARMACIAS NISSEI' : 'FARMÁCIAS E DROGARIAS NISSEI LTDA',
        'CIA BEAL DE ALIMENTOS' : 'CIA BEAL DE ALIMENTOS',
        'NATUCLIN' : 'NATUCLIN COMÉRCIO DE PRODUTOS NATURAIS E MANUFATURADOS LTDA',
        'AUTOSHOPPING LINHA VERDE' : 'ASSOCIACAO DOS LOJISTAS DO AUTO SHOPPING LINHA VERDE',
        'AUTO SHOPPING LINHA VERDE' : 'ASSOCIACAO DOS LOJISTAS DO AUTO SHOPPING LINHA VERDE',
        'SUPERMERCADO JACOMAR' : 'SUPERMERCADO JACOMAR LTDA',
        'INST ODONTOLOGICO DR CHRISTIAN ANDRADE' : 'C & R ODONTOLOGIA LTDA',
        'O BOTICARIO' : 'BOTICARIO PRODUTOS DE BELEZA LTDA',
        'FG EMPREENDIMENTOS' : 'FG BRAZIL HOLDING LTDA',
        'SERCOMTEL TELECOMUNICACOES' : 'LIGGA TELECOM',
        'BIRD EVENTOS' : 'MINASI & CIA LTDA',
        'UNICESUMAR' : 'CENTRO DE ENSINO SUPERIOR MARINGA LTDA - UNICESUMAR',
        'COAMO AGROINDUSTRIAL COOP' : 'COAMO COOPERATIVA AGROINDUSTRIAL',
        'BMG FOODS' : 'BMG FOODS IMPORTAÇÃO E EXPORTAÇÃO LTDA',
        'MOVEIS GAZIN' : 'GAZIN INDUSTRIA E COMERCIO DE MOVEIS E ELETRODOMESTICOS LTDA',
        'PREVER SERVICOS POSTUMOS' : 'PREVER SERVIÇOS POSTUMOS LTDA',
        'CIA SULAMERICANA DE DISTR' : 'COMPANHIA SULAMERICANA DE DISTRIBUICAO - CSD',
        'IGUARACU EMPRE IMOBILIARIOS' : 'IGUARACU EMPREENDIMENTOS IMOBILIARIOS LTDA - ME',
        'LOJA DA CASA' : 'M S G PISOS CERAMICOS LTDA',
        'ORAL TIME SERVICOS ODONTOLOGICOS' : 'ORAL TIME SERIÇOS ODONTOLOGICOS LTDA',
        'LOJAS ALVORADA' : 'LOJAS ALVORADA LTDA - ME',
        'IMPACTO FRANQUIAS' : 'J A L IMPACTO MARINGA COMERCIO DE PECAS E PNEUS EIRELI',
        'PONTO TRACK RASTREAMENTO E LOGISTICA' : 'PONTO TRACK RASTREAMENTO E LOGISTICA LTDA ME',
        'SATTRACK RASTREAMENTO E LOGISTICA' : 'SATTRACK RASTREAMENTO E LOGISTICA LTDA - ME',
        'SUPERMERCADO TONHAO' : 'SUPERMERCADO TONHAO LTDA ME',
        'VERONA SUPERMERCADOS' : 'SANCHES E VECCHIATE LTDA',
        'SERCOMTEL TELECOMUNICACOES' : 'SERCOMTEL S/A TELECOMUNICAÇÕES',
        'MOLICENTER' : 'MOLICENTER SUPERMERCADO LTDA',
        'SUPER PUPPO SUPERMERCADOS' : 'SUPER PUPPO SUPERMERCADOS LTDA',
        'SUPERMERCADO IRANI' : 'SUPERMERCADOS IRANI LTDA',
        'LOVAT HYUNDAI' : 'LOVAT VEICULOS S/A',
        'LOVAT HYUNDAI' : 'LOVAT VEÍCULOS LTDA',
        'SUPERDIA ATACADO' : 'ITALO SUPERMERCADOS LTDA',
        'HOSPITAL DE OLHOS DE CASCAVEL' : 'HOSPITAL DE OLHOS CENTRO OFTALMOLOGICO DE CASCAVEL LTDA',
        'HIPERZOO PET SHOP' : 'HIPERZOO PET SHOP LTDA',
        'ALTHAIA' : 'ALTHAIA S.A. INDUSTRIA FARMACEUTICA',
        'ALIANCA MOVEIS' : 'COMERCIO DISTRIBUIDOR DE MOVEIS LTDA',
        'INST DOS OCULOS PR' : 'INSTITUTO DOS OCULOS FRANCHISING EIRELLI - ME',
        'SHOPPING DOS ENXOVAIS' : 'BMQ ENXOVAIS EIRELI',
        'FOLHA DE LONDRINA' : 'EDITORA E GRAFICA PARANA PRESS S.A.',
        'MOVEIS BRASILIA' : 'COMERCIAL DE MÓVEIS BRASÍLIA LTDA',
        'Y AGITA COM DE CALCADOS' : 'Y AGITA COMÉRCIO DE CALÇADOS LTDA',
        'RECANTO CATARATAS THERMAS RESORT E CONVENTION' : 'RECANTO CATARATAS HOTEL E CONVENTION LTDA',
        'SAFEAGRO' : 'SAFEAGRO AGROCIENCIA LTDA',
        'C VALE COOP AGROINDUSTRIAL' : 'C.VALE COOPERATIVA AGROINDUSTRIAL',
        'SOLUCAO FINANCEIRA' : 'SOLUCAO FINANCEIRA - SERVICOS DE RECUPERACAO DE CREDITO EIRELI',
        'XBRI IMP BRASILEIRA DE PNEUS' : 'EAA - TOLEDO COMERCIO DE PECAS E PNEUS LTDA',
        'EMPORIO SANTA MARIA' : 'EMPORIO SANTA MARIA ENTRETENIMENTO LTDA',
        'EMPORIO CASA DE EVENTOS' : 'EMPORIO SANTA MARIA ENTRETENIMENTO LTDA',
        'ODONTO TOP HOSPITAL DO DENTE PATO BRANCO' : 'ODONTOTOP TOLEDO LTDA',
        'COLCHOES NIPPON BRASIL' : 'ADILSON PEREIRA',
        'OTICA POPULAR CONCEITO' : 'OTICA POPULAR TRAMONTIN E BIERGER LTDA',
        'JD HOME CENTER' : 'MOVISTAR COMERCIO DE MATERIAL DE CONSTRUCAO LTDA',
        'CONSTRUCAL' : 'CONSTRUCAL MATERIAIS DE CONSTRUCAO LTDA',
        'LOJAS KELLI' : 'LJS CONFECÇÕES LTDA',
        'BUENAO IND E COM DE CONFECCOES' : 'LB BUENAO COMERCIO DE CONFECCOES LTDA.',
        'MOTOPECAS TOLEDO' : 'DISTRIBUIDORA DE MOTO PECAS TOLEDO LTDA',
        'LOJA BOM SUCESSO CASCAVEL' : 'MARCHIORI E BRUGNOLI BRUGNOLI LTDA ME',
        'LOJA BOM SUCESSO CHAPECO' : 'MARCHIORI E BRUGNOLI BRUGNOLI LTDA ME',
        'UNIMED COSTA OESTE' : 'UNIMED COSTA OESTE COOPERATIVA DE TRABALHO MEDICO',
        'PANORAMA HOME CENTER' : 'PANORAMA MATERIAIS DE CONSTRUCAO LTDA',
        'COMBATE DINIZ' : 'COMBATE DINIZ MOVEIS E ELETRO LTDA',
        'GHELERE TRANSPORTES' : 'GHELERE TRANSPORTES LTDA',
        'MORBEH SUPERMERCADOS' : 'MORBEH SUPERMERCADOS LTDA',
        'SUPERGASBRAS ENERGIA' : 'SUPER-TOLEDO GAS LTDA',
        'TOLE POCOS ARTESIANOS' : 'TOLE POCOS ARTESIANOS LTDA',
        'COOPAVEL' : 'COOPAVEL COOPERATIVA AGROINDUSTRIAL',
        'PRIMATO' : 'PRIMATO COOPERATIVA AGROINDUSTRIAL',
        'SHOPPING CHINA COM UTILIDADES DOMESTICAS' : 'IP COM COMERCIO DE EQUIPAMENTOS DE TELEFONIA LTDA',
        'HOSPITAL POLICLINICA CASCAVEL' : 'HOSPITAL POLICLINICA CASCAVEL S.A',
        'CENTRAL TINTAS' : 'CENTRAL TINTAS COMERCIO DE TINTAS LTDA',
        'UNITOM' : 'UNITOM UNIDADE DE DIAGNOSTICO POR IMAGEM S/S LTDA',
        'DESTAK ODONTOLOGIA' : 'DESTAK ODONTOLOGIA LTDA',
        'AGUA MINERAL ITAIPU' : 'EMPRESA DE AGUA MINERAL ITAIPU LTDA',
        'ALIMENTOS ZAELI' : 'ALIMENTOS ZAELI LTDA',
        'R C MACHADO' : 'R.C.MACHADO',
        'ARMAZEM DA MARIA' : 'COMERCIAL DESTRO LTDA',
        'BLUEFIT ACADEMIAS GINASTICA PARTICIPACOES' : 'WR KRUGER E HUBNER ACADEMIA LTDA',
        'CBS EMPREENDIMENTOS IMOBILIARIOS' : 'CBS - EMPREENDIMENTOS IMOBILIARIOS SPE LTDA',
        'UMUPREV' : 'UMUPREV PLANO DE ASSISTENCIA FAMILIAR LTDA',
        'ORAL UNIC IMPLANTES' : 'ORAL UNIC ODONTOLOGIA TOLEDO LTDA',
        'MARCELO CALCADOS E ACESSORIOS' : 'MARCELO CALCADOS E ACESSORIOS EIRELI',
        'IND E COM DE LATICINIOS PEREIRA' : 'INDUSTRIA E COMERCIO DE LATICINIOS PEREIRA LTDA',
        'ENGIE BRASIL ENERGIA' : 'ENGIE BRASIL ENERGIA S A',
        'CIA PARANAENSE DE ENERGIA (GEP)' : 'COMPANHIA PARANAENSE DE ENERGIA',
        'SUPERMERCADO PLANALTO' : 'J MARTINS SUPERMERCADOS PLANALTO LTDA',
        'CARRO VENDIDO' : 'CV MARINGÁ TECNOLOGIA EM SERVIÇOS DE VENDA LTDA',
        'UNINGA' : 'UNINGA - UNID.ENSINO SUP.INGA SC LT',
        'SHIMIZU IMOVEIS' : 'SHIMIZU IMOVEIS LTDA',
        'MOVEIS SAO CARLOS' : 'MSC COMERCIO DE MOVEIS LTDA',
        'CHANSON VEICULOS' : 'CHANSON',
        'MIYAMOTO OBARA E CIA' : 'OBARA MIYAMOTO & CIA LTDA',
        'GCA DISTR COML DE ALIMENTOS' : 'GCA - DISTRIBUIDORA COMERCIAL DE ALIMENTOS LTDA',
        'SANEPAR SANEAMENT PARANA (GEP)' : 'COMPANHIA DE SANEAMENTO DO PARANA - SANEPAR',
        'ESTRE AMBIENTAL' : 'ESTRE AMBIENTAL S/A EM RECUPERAÇÃO JUDICIAL',
        'ADEMICON ADM DE CONSORCIOS' : 'ADEMICON ADMINISTRADORA DE CONSORCIOS S/A',
        'MINIPRECO' : 'JUMBO COMERCIO DE UTILIDADES LTDA',
        'ITAIPU BINACIONAL (GFP)' : 'ITAIPU',
        'BERTOLDO E PELEGRINO' : 'BERTOLDO & PELEGRINO LTDA - ME',
        'BATISTA & IZEPE LTDA' : 'REDE BOM DIA SUPERMERCADOS',
        'UNINTER EDUCACIONAL' : 'UNINTER EDUCACIONAL S/A',
        'VITAO ALIMENTOS' : 'VITAO ALIMENTOS LTDA',
        'SUPERMERCADOS RIO VERDE' : 'GUSTO E HENRI SUPERMERCADOS LTDA',
        '' : '',
        
        #PREFEITURAS:
        'PREF MUN CURITIBA (GMP)' : 'MUNICIPIO DE CURITIBA',
        'PREF MUN GUARATUBA (GMP)' : 'MUNICIPIO DE GUARATUBA',
        'PREF MUN FAZEN RIO GRANDE (GMP)' : 'PREFEITURA MUNICIPAL FAZENDA RIO GRANDE',
        'PREF MUN ARAUCARIA (GMP)' : 'PREFEITURA DO MUNICIPIO DE ARAUCARIA',
        'PREF MUN MARINGA (GMP)' : 'PREFEITURA DO MUNICIPIO DE MARINGA',
        'PREF MUN LONDRINA (GMP)' : 'PREFEITURA DO MUNICIPIO DE LONDRINA',
        'PREF MUN CASCAVEL (GMP)' : 'MUNICIPIO DE CASCAVEL',
        'PREF MUN TOLEDO (GMP)' : 'MUNICIPIO DE TOLEDO',
        'PREF MUN FOZ DO IGUACU (GMP)' : 'FOZ DO IGUACU PREFEITURA',
        'SECR MUN SAUDE FRANCISCO BELTRAO (GMP)' : 'MUNICIPIO DE FRANCISCO BELTRAO',
        'PREF MUN PONTA GROSSA (GMP)' : 'PREFEITURA MUNICIPAL DE PONTA GROSSA',
    }

    def update_valor_liquido(row, client_mapping, Bdf):
        # Define a mapping for praca, coverage, and their corresponding filters
        praca_coverage_filters = {
            #OESTE coverage
            ('MARINGA', 'OESTE'): {'Exibição_contains': 'MARINGÁ', 'Emp. Venda': 26},
            ('CASCAVEL', 'OESTE'): {'Exibição_contains': 'TOLEDO', 'Emp. Venda': 26},
            ('CURITIBA', 'OESTE'): {'Exibição_contains': 'CURITIBA', 'Emp. Venda': 26},
            ('LONDRINA', 'OESTE'): {'Exibição_contains': 'LONDRINA', 'Emp. Venda': 26},
            #OESTE REDE coverage
            #('MARINGA', 'OESTE'): {'Exibição_contains': 'REDE', 'Emp. Venda': 26},
            #('LONDRINA', 'OESTE'): {'Exibição_contains': 'REDE', 'Emp. Venda': 26},
            #('CASCAVEL', 'OESTE'): {'Exibição_contains': 'REDE', 'Emp. Venda': 26},
            #('CURITIBA', 'OESTE'): {'Exibição_contains': 'REDE', 'Emp. Venda': 26},

            #LON coverage now
            ('LONDRINA', 'LON'): {'Exibição_contains': 'LONDRINA', 'Emp. Venda': 25},
            ('CURITIBA', 'LON'): {'Exibição_contains': 'CURITIBA', 'Emp. Venda': 25},
            ('MARINGA', 'LON'): {'Exibição_contains': 'MARINGÁ', 'Emp. Venda': 25},
            ('CASCAVEL', 'LON'): {'Exibição_contains': 'TOLEDO', 'Emp. Venda': 25},
            #LON REDE coverage
            #('LONDRINA', 'LON'): {'Exibição_contains': 'REDE', 'Emp. Venda': 25},
            #('CURITIBA', 'LON'): {'Exibição_contains': 'REDE', 'Emp. Venda': 25},
            #('MARINGA', 'LON'): {'Exibição_contains': 'REDE', 'Emp. Venda': 25},
            #('CASCAVEL', 'LON'): {'Exibição_contains': 'REDE', 'Emp. Venda': 25},

            #MAR coverage now
            ('MARINGA', 'MAR'): {'Exibição_contains': 'MARINGÁ', 'Emp. Venda': 24},
            ('CURITIBA', 'MAR'): {'Exibição_contains': 'CURITIBA', 'Emp. Venda': 24},
            ('LONDRINA', 'MAR'): {'Exibição_contains': 'LONDRINA', 'Emp. Venda': 24},
            ('CASCAVEL', 'MAR'): {'Exibição_contains': 'TOLEDO', 'Emp. Venda': 24},
            #MAR REDE coverage
            #('MARINGA', 'MAR'): {'Exibição_contains': 'REDE', 'Emp. Venda': 24},
            #('CURITIBA', 'MAR'): {'Exibição_contains': 'REDE', 'Emp. Venda': 24},
            #('LONDRINA', 'MAR'): {'Exibição_contains': 'REDE', 'Emp. Venda': 24},
            #('CASCAVEL', 'MAR'): {'Exibição_contains': 'REDE', 'Emp. Venda': 24},

            #CTBA coverage now
            ('CURITIBA', 'CTBA'): {'Exibição_contains': 'CURITIBA', 'Emp. Venda': 23},
            ('LONDRINA', 'CTBA'): {'Exibição_contains': 'LONDRINA', 'Emp. Venda': 23},
            ('MARINGA', 'CTBA'): {'Exibição_contains': 'MARINGÁ', 'Emp. Venda': 23},
            ('CASCAVEL', 'CTBA'): {'Exibição_contains': 'TOLEDO', 'Emp. Venda': 23},
            #CTBA REDE coverage
            #('CURITIBA', 'CTBA'): {'Exibição_contains': 'REDE', 'Emp. Venda': 23},
            #('LONDRINA', 'CTBA'): {'Exibição_contains': 'REDE', 'Emp. Venda': 23},
            #('MARINGA', 'CTBA'): {'Exibição_contains': 'REDE', 'Emp. Venda': 23},
            #('CASCAVEL', 'CTBA'): {'Exibição_contains': 'REDE', 'Emp. Venda': 23},
        }

        # Extract information from the row
        client_name_df = row['Anunciante'].upper()
        praca = row['Praça'].upper()
        emissora = row['Emissora TV'].upper()
        coverage = row['Cobertura'].upper()

        # Resolve the client name using the mapping or default to the original name
        client_name_bdf = client_mapping.get(client_name_df, client_name_df)

        # Initialize price_per_row to the current value (default case)
        price_per_row = row['Valor Líquido Projetado']

        # Check for NaN or 0 in 'Valor Líquido Projetado'
        if pd.isna(price_per_row) or price_per_row == 0:
            # Lookup the specific filters for the current praca and coverage
            filters = praca_coverage_filters.get((praca, coverage))

            # If specific filters are found, proceed with the calculation
            if filters and emissora == 'RECORD TV':
                # Apply filters to Bdf
                filtered_bdf_rows = Bdf[
                    (Bdf['Cliente'] == client_name_bdf) &
                    (Bdf['Exibição'].str.contains(filters['Exibição_contains'])) &
                    (Bdf['Emp. Venda'] == filters['Emp. Venda'])
                ]

                # Calculate the total price from filtered Bdf rows
                total_price = filtered_bdf_rows['Valor Líquido'].sum()

                # Define additional filters for the main DataFrame based on praca and coverage
                additional_filters = (
                    (df['Praça'].str.upper() == praca) &
                    (df['Cobertura'].str.upper() == coverage) &
                    (df['Emissora TV'].str.upper() == emissora)
                )

                # Count matching rows in the main DataFrame
                num_filtered_rows_in_df = df[(df['Anunciante'].str.upper() == client_name_df) & additional_filters].shape[0]

                # Avoid division by zero
                if num_filtered_rows_in_df > 0:
                    # Calculate and return the price per row
                    price_per_row = total_price / num_filtered_rows_in_df

        return price_per_row

    # Assuming 'client_mapping' and 'Bdf' are defined elsewhere in your script

    # Update 'Valor Líquido Projetado' column
    df['Valor Líquido Projetado'] = df.apply(lambda row: update_valor_liquido(row, client_mapping, Bdf), axis=1)

    print(df)

    # # Let's make an excel file from this bad boy
        
    ignore = ['PREF SEDE', 'GOVERNO', 'ASSEMBLEIA']

    curitiba_zero = df.loc[
        (df['Cobertura'] == 'CTBA') &
        (df['Emissora TV'].str.contains('RECORD')) &
        (~df['Mercado'].isin(ignore)) &
        (df['Valor Líquido Projetado'] == 0)
    ]['Valor Líquido Projetado'].count()
    status_elem.print(f'CWB - A quantidade de linhas não preenchida é: {curitiba_zero}')

    curitiba_value = df.loc[
        (df['Cobertura'] == 'CTBA') &
        (df['Emissora TV'].str.contains('RECORD')) &
        (~df['Mercado'].isin(ignore)) &
        (df['Valor Líquido Projetado'] != 0)
    ]['Valor Líquido Projetado'].sum()

    try:
        valorbasCWB = int(basket_values[0])
    except ValueError:
        print("O valor digitado não é um número inteiro válido.")
        valorbasCWB = 0  # or handle this as needed
    status_elem.print(f'O valor de Curitiba é:{valorbasCWB}')

    # Safety check to avoid division by zero
    if curitiba_zero != 0:
        CWBdiff = (valorbasCWB - curitiba_value) / curitiba_zero
    else:
        print("No rows match the criteria for adjustment. Check your data and criteria.")
        CWBdiff = 0

    # Only proceed with the update if CWBdiff is calculated successfully (i.e., not dividing by zero)
    if CWBdiff:
        df.loc[
            (df['Cobertura'] == 'CTBA') &
            (df['Emissora TV'].str.contains('RECORD')) &
            (~df['Mercado'].isin(ignore)) &
            (df['Valor Líquido Projetado'] == 0),
            'Valor Líquido Projetado'
        ] = CWBdiff
    else:
        print("No update was made to 'Valor Líquido Projetado'.")

    maringa_zero = df.loc[
        (df['Cobertura'] == 'MAR') &
        (df['Emissora TV'].str.contains('RECORD')) &
        (~df['Mercado'].isin(ignore)) &
        (df['Valor Líquido Projetado'] == 0)
    ]['Valor Líquido Projetado'].count()
    status_elem.print(f'MAR - A quantidade de linhas não preenchida é: {maringa_zero}')

    maringa_value = df.loc[
        (df['Cobertura'] == 'MAR') &
        (df['Emissora TV'].str.contains('RECORD')) &
        (~df['Mercado'].isin(ignore)) &
        (df['Valor Líquido Projetado'] != 0)
    ]['Valor Líquido Projetado'].sum()

    try:
        valorbasMAR = int(basket_values[1])
    except ValueError:
        print("O valor digitado não é um número inteiro válido.")
        valorbasMAR = 0  # or handle this as needed
    status_elem.print(f'O valor de Maringá é:{valorbasMAR}')


    # Safety check to avoid division by zero
    if maringa_zero != 0:
        MARdiff = (valorbasMAR - maringa_value) / maringa_zero
    else:
        print("No rows match the criteria for adjustment. Check your data and criteria.")
        MARdiff = 0

    # Only proceed with the update if CWBdiff is calculated successfully (i.e., not dividing by zero)
    if MARdiff:
        df.loc[
            (df['Cobertura'] == 'MAR') &
            (df['Emissora TV'].str.contains('RECORD')) &
            (~df['Mercado'].isin(ignore)) &
            (df['Valor Líquido Projetado'] == 0),
            'Valor Líquido Projetado'
        ] = MARdiff
    else:
        print("No update was made to 'Valor Líquido Projetado'.")


    londrina_zero = df.loc[
        (df['Cobertura'] == 'LON') &
        (df['Emissora TV'].str.contains('RECORD')) &
        (~df['Mercado'].isin(ignore)) &
        (df['Valor Líquido Projetado'] == 0)
    ]['Valor Líquido Projetado'].count()
    status_elem.print(f'LON - A quantidade de linhas não preenchida é: {londrina_zero}')

    londrina_value = df.loc[
        (df['Cobertura'] == 'LON') &
        (df['Emissora TV'].str.contains('RECORD')) &
        (~df['Mercado'].isin(ignore)) &
        (df['Valor Líquido Projetado'] != 0)
    ]['Valor Líquido Projetado'].sum()

    try:
        valorbasLON = int(basket_values[2])
    except ValueError:
        print("O valor digitado não é um número inteiro válido.")
        valorbasLON = 0  # or handle this as needed
    status_elem.print(f'O valor de Londrina é:{valorbasLON}')

    # Safety check to avoid division by zero
    if londrina_zero != 0:
        LONdiff = (valorbasLON - londrina_value) / londrina_zero
    else:
        print("No rows match the criteria for adjustment. Check your data and criteria.")
        LONdiff = 0

    # Only proceed with the update if CWBdiff is calculated successfully (i.e., not dividing by zero)
    if LONdiff:
        df.loc[
            (df['Cobertura'] == 'LON') &
            (df['Emissora TV'].str.contains('RECORD')) &
            (~df['Mercado'].isin(ignore)) &
            (df['Valor Líquido Projetado'] == 0),
            'Valor Líquido Projetado'
        ] = LONdiff
    else:
        print("No update was made to 'Valor Líquido Projetado'.")

    oeste_zero = df.loc[
        (df['Cobertura'] == 'OESTE') &
        (df['Emissora TV'].str.contains('RECORD')) &
        (df['Valor Líquido Projetado'] == 0) &
        (~df['Mercado'].isin(ignore))
    ]['Valor Líquido Projetado'].count()
    status_elem.print(f'OESTE - A quantidade de linhas não preenchida é: {oeste_zero}')

    oeste_value = df.loc[
        (df['Cobertura'] == 'OESTE') &
        (df['Emissora TV'].str.contains('RECORD')) &
        (df['Valor Líquido Projetado'] != 0) &
        (~df['Mercado'].isin(ignore))
    ]['Valor Líquido Projetado'].sum()

    try:
        valorbasOESTE = int(basket_values[3])
    except ValueError:
        print("O valor digitado não é um número inteiro válido.")
        valorbasOESTE = 0  # or handle this as needed
    status_elem.print(f'O valor de Oeste é:{valorbasOESTE}')

    # Safety check to avoid division by zero
    if oeste_zero != 0:
        OESTEdiff = (valorbasOESTE - oeste_value) / oeste_zero
    else:
        print("No rows match the criteria for adjustment. Check your data and criteria.")
        OESTEdiff = 0

    # Only proceed with the update if CWBdiff is calculated successfully (i.e., not dividing by zero)
    if OESTEdiff:
        df.loc[
            (df['Cobertura'] == 'OESTE') &
            (df['Emissora TV'].str.contains('RECORD')) &
            (~df['Mercado'].isin(ignore)) &
            (df['Valor Líquido Projetado'] == 0),
            'Valor Líquido Projetado'
        ] = OESTEdiff
    else:
        print("No update was made to 'Valor Líquido Projetado'.")

    # Create or get the file path
    file_path = create_file_path(save_folder, month_name, result_name)

    # Write or append the DataFrame to the Excel file
    write_to_excel(df, file_old, file_path, append_to_file)
    
    time.sleep(2)

    # Attempt to open the Excel file
    status_message = open_file(file_path)
    status_elem.print(status_message)
    time.sleep(2)

    try:
    # Open the Excel file using the default application
        subprocess.Popen(["start", file_path], shell=True)  # On Windows

        status_elem.print(f"Abrindo {file_path} no aplicativo padrao... Sucesso!")
    except FileNotFoundError:
            status_elem.print(f"Error: File '{file_path}' not found.")
    except Exception as e:
            status_elem.print(f"An error occurred: {e}")

    status_elem.print('bip bip bop terminamos!')

# Layout for the initial file selection window
layout_file_selection = [
    [sg.Frame('Selecione os Arquivos Excel:', [
        [sg.Text('MonitorFlex'), sg.InputText(key='_FILE1_'), sg.FileBrowse()],
        #[sg.Text('BASQUETE'), sg.InputText(key='_FILE2_'), sg.FileBrowse()],
        [sg.Text('COBERTURA'), sg.InputText(key='_FILE3_'), sg.FileBrowse()],
    ])],
    [sg.Frame('Configurações de Salvamento:', [
        [sg.Text('Selecione a Pasta de Salvamento:'), sg.InputText(key='_SAVE_FOLDER_', disabled=False), sg.FolderBrowse(key='_SAVE_FOLDER_BROWSE_', disabled=False)],
        [sg.Checkbox('Adicionar ao arquivo existente?', default=False, key='APPEND', enable_events=True),
         sg.InputText('', key='_APPEND_FILE_', disabled=True), sg.FileBrowse(button_text='Procurar', key='_APPEND_FILE_BROWSE_', target='_APPEND_FILE_', disabled=True)],
        [sg.Text('Digite o nome do resultado:'), sg.InputText(key='resul')]
    ])],
    [sg.Frame('Informações Adicionais:', [
        [sg.Text('Digite o Número do Mês'), sg.InputText(key='month')],
        [sg.Text('Digite o Ano Desejado'), sg.InputText(key='year')],
        [sg.Text('Digite o valor total de Curitiba:'), sg.InputText(key='CWB')],
        [sg.Text('Digite o valor total de Maringa:'), sg.InputText(key='MAR')],
        [sg.Text('Digite o valor total de Londrina:'), sg.InputText(key='LON')],
        [sg.Text('Digite o valor total de Cascavel:'), sg.InputText(key='OES')],
    ])],
    [sg.Button('Enviar')]
]

# Layout for the status window
layout_status = [
    [sg.Text('Status:')],
    [sg.Multiline('', size=(60, 15), key='_STATUS_', autoscroll=True)],
    [sg.Button('Fechar')]
]

# Create the initial file selection window
window_file_selection = sg.Window('File Selection', layout_file_selection, ttk_theme=ttk_style)

# Event loop to capture user inputs for file selection
while True:
    event, values = window_file_selection.read()

    if event == sg.WIN_CLOSED:
        break

    # Handle checkbox toggle
    elif event == 'APPEND':
        append_to_file = values['APPEND']  # Get the boolean value from the checkbox

        # Enable or disable the file browse field based on checkbox state
        window_file_selection['_APPEND_FILE_'].update(disabled=not values['APPEND'])
        window_file_selection['_APPEND_FILE_BROWSE_'].update(disabled=not values['APPEND'])
        window_file_selection['_SAVE_FOLDER_'].update(disabled= values['APPEND'])
        window_file_selection['_SAVE_FOLDER_BROWSE_'].update(disabled= values['APPEND'])

    elif event == 'Enviar':

        save_folder = values['_SAVE_FOLDER_'] 

        month_name = get_month_name(int(values['month']))  # Assuming you have a `get_month_name` function
        result_name = values['resul']
        append_to_file = values['APPEND']
        file_old = values['_APPEND_FILE_']
        year_sheet = values['year']
        

        file_paths = [
            values['_FILE1_'],
            #values['_FILE2_'],
            values['_FILE3_'],
            values['month'],
            values['resul'],
        ]
        basket_values = [
            int(values['CWB']),
            int(values['MAR']),
            int(values['LON']),
            int(values['OES']),
        ]
        window_file_selection.close()

        layout_status = [
            [sg.Text('Status:')],
            [sg.Multiline('', size=(60, 15), key='_STATUS_', autoscroll=True)],
            [sg.Button('Fechar')]
        ]

        window_status = sg.Window('Status Window', layout_status, finalize=True)

        thread = threading.Thread(target=execute_code, args=(file_paths, window_status, append_to_file, file_old))
        thread.daemon = True
        thread.start()

        while True:
            event_status, values_status = window_status.read()
            if event_status == sg.WIN_CLOSED or event_status == 'Fechar':
                break
        window_status.close()
        break

sg.popup('Programa Finalizado!')